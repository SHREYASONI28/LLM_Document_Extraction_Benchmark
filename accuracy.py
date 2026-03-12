import pandas as pd
import json
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher
import argparse
import re

INPUT_FILE = "benchmark_output.xlsx"
OUTPUT_FILE = "benchmark_accuracy.xlsx"

# ---------------- CLI INPUT ----------------
parser = argparse.ArgumentParser()
parser.add_argument("--config", required=False)
parser.add_argument("--prompt", required=False)

args = parser.parse_args()

# ---------------- LOAD CONFIG ----------------
if args.config:
    with open(args.config, "r") as f:
        config = json.load(f)
    FIELDS = config["fields"]

elif args.prompt:
    match = re.search(r"extract\s+(.*)", args.prompt.lower())
    if match:
        FIELDS = [f.strip() for f in match.group(1).split(",")]
    else:
        raise ValueError("Prompt must specify fields like: Extract name, email, phone")

else:
    raise ValueError("Provide config or prompt")

# ---------------- NORMALIZATION ----------------
def normalize(text):
    text = str(text).lower().strip()
    text = text.replace(",", "")
    text = text.replace("₹", "")
    text = text.replace("$", "")
    text = text.replace("inr", "")
    return text
    
#Similarity Function
def is_match(a, b, threshold=0.95):
    ratio = SequenceMatcher(None, a, b).ratio()
    
    length_diff = abs(len(a) - len(b)) / max(len(a), 1)

    return ratio >= threshold and length_diff < 0.2
# ---------------- MAIN ----------------
xls = pd.ExcelFile(INPUT_FILE)
writer = pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl")

summary = []

for sheet in xls.sheet_names:

    if sheet == "Final_Summary":
        continue

    df = pd.read_excel(INPUT_FILE, sheet_name=sheet)

    models = df.columns[2:]

    metrics = {
        model: {"TP": 0, "FP": 0, "FN": 0}
        for model in models
    }

    execution_times = {}

    valid_rows = 0

    for i in range(len(df)):

        field = str(df.loc[i, "Field"])

        # Capture execution time row
        if field.startswith("TOTAL_MODEL_EXECUTION_TIME"):
            for model in models:
                execution_times[model] = df.loc[i, model]
            continue

        
        gt_value = normalize("" if pd.isna(df.loc[i, "Ground Truth"]) else df.loc[i, "Ground Truth"])
        # Skip rows with empty ground truth
        if gt_value == "":
            continue
        valid_rows += 1
        for model in models:

            pred_value = normalize("" if pd.isna(df.loc[i, model]) else df.loc[i, model])
            # check if GT appears inside prediction
            if gt_value in pred_value:
                metrics[model]["TP"] += 1

            elif is_match(gt_value, pred_value):
                metrics[model]["TP"] += 1

            elif pred_value == "":
                metrics[model]["FN"] += 1

            else:
                metrics[model]["FP"] += 1
    # ---- Calculate Accuracy ----
    results = {}

    for model in models:

        TP = metrics[model]["TP"]
        FP = metrics[model]["FP"]
        FN = metrics[model]["FN"]

        precision = TP / (TP + FP) if (TP + FP) else 0
        recall = TP / (TP + FN) if (TP + FN) else 0

        f1 = (
            2 * precision * recall / (precision + recall)
            if (precision + recall)
            else 0
        )

        accuracy = TP / valid_rows if valid_rows else 0

        results[model] = {
            "accuracy": round(accuracy * 100, 2),
            "precision": round(precision * 100, 2),
            "recall": round(recall * 100, 2),
            "f1": round(f1 * 100, 2)
        }
        

    # ---- Write Original Sheet As It Is ----
    df.to_excel(writer, sheet_name=sheet, index=False)

    ws = writer.sheets[sheet]

    # ---------------- FORMATTING ----------------
    bold_font = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header styling
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Data styling
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = wrap_align

    # Bold Field column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=1):
        for cell in row:
            cell.font = bold_font

    # Bold execution time row
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if str(row[0].value).startswith("TOTAL_MODEL_EXECUTION_TIME"):
            for cell in row:
                cell.font = bold_font
                cell.alignment = center_align

    # Auto column width
    for col in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

    # ---- Add to Final Summary ----
    summary.append([sheet, "Accuracy"] + [results[m]["accuracy"] for m in models])
    summary.append(["", "Precision"] + [results[m]["precision"] for m in models])
    summary.append(["", "Recall"] + [results[m]["recall"] for m in models])
    summary.append(["", "F1 Score"] + [results[m]["f1"] for m in models])

# ---------------- FINAL SUMMARY SHEET ----------------

columns = ["Document", "Metric"] + list(models)

summary_df = pd.DataFrame(summary, columns=columns)
summary_df.to_excel(writer, sheet_name="Final_Summary", index=False)

ws = writer.sheets["Final_Summary"]

# Styling Final Summary
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                        min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Bold document column
for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                        min_col=1, max_col=1):
    for cell in row:
        cell.font = bold_font
# Force text format for Document column
for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                        min_col=1, max_col=1):
    for cell in row:
        cell.number_format = "@"
# Auto width
for col in range(1, ws.max_column + 1):
    max_length = 0
    col_letter = get_column_letter(col)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=col, max_col=col):
        for cell in row:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

    ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

writer.close()

print("\nDONE → benchmark_accuracy.xlsx created")