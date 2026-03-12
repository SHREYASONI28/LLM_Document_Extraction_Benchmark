import os
import time
import subprocess
import pandas as pd
from pypdf import PdfReader
from docx import Document
import openpyxl
import json
import re
import argparse
from dotenv import load_dotenv
from openai import AzureOpenAI

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# -------- LOAD AZURE ENV VARIABLES --------
load_dotenv()

AZURE_API_KEY = os.getenv("AZURE_OPENAI_API_KEY")
AZURE_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")

AZURE_API_VERSION_GPT = os.getenv("AZURE_API_VERSION_GPT")
AZURE_API_VERSION_LLAMA = os.getenv("AZURE_API_VERSION_LLAMA")

# ---------------- CLI INPUT ----------------
parser = argparse.ArgumentParser()

parser.add_argument(
    "--input",
    required=True,
    help="PDF,TXT , DOCX or XLSX file path OR comma separated list of files"
)
parser.add_argument(
    "--config",
    required=False,
    help="Path to config JSON file"
)

parser.add_argument(
    "--prompt",
    required=False,
    help="Direct extraction prompt"
)
args = parser.parse_args()

input_files = args.input.split(",")

# ---------------- LOAD CONFIG OR PROMPT ----------------

if args.config:
    with open(args.config, "r") as f:
        config = json.load(f)

    FIELDS = config["fields"]
    USER_PROMPT = config["prompt"]
elif args.prompt:
    USER_PROMPT = args.prompt

    # Extract fields from prompt automatically
    match = re.search(r"extract\s+(.*)", args.prompt.lower())

    if match:
        FIELDS = [f.strip() for f in match.group(1).split(",")]
    else:
        raise ValueError("Prompt must specify fields like: Extract name, email, phone")
else:
    raise ValueError("Provide either --config or --prompt")

OUTPUT_FILE = "benchmark_output.xlsx"



# ---------------- PDF TEXT EXTRACTION ----------------
def pdf_text(path):
    reader = PdfReader(path)
    text = ""

    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            text += page_text + "\n"
    # Remove problematic characters
    text = text.replace("■", "")
    text = text.encode("ascii", "ignore").decode()

    return text.strip()

# ---------------- TXT TEXT EXTRACTION ----------------
def txt_text(path):
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()

    text = text.replace("■", "")
    text = text.encode("ascii", "ignore").decode()

    return text.strip()

# ---------------- DOCX TEXT EXTRACTION ----------------
def docx_text(path):
    doc = Document(path)
    text = ""

    for para in doc.paragraphs:
        text += para.text + "\n"

    text = text.replace("■", "")
    text = text.encode("ascii", "ignore").decode()

    return text.strip()

# ---------------- EXCEL TEXT EXTRACTION ----------------
def excel_text(path):
    wb = openpyxl.load_workbook(path)
    text = ""

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    text += str(cell) + " "

    text = text.replace("■", "")
    text = text.encode("ascii", "ignore").decode()

    return text.strip()

# ---------------- LLM CALL ----------------
def ollama_json(model, text):

    # build dynamic JSON template
    json_template = "{\n"
    for field in FIELDS:
        json_template += f' "{field}": "",\n'
    json_template = json_template.rstrip(",\n") + "\n}"

    prompt = f"""
{USER_PROMPT}

If not found, use empty string "".

Return ONLY ONE valid JSON object.
No explanation.

{json_template}

Document:
{text}
"""

    try:
        r = subprocess.run(
            ["ollama", "run", model, "--verbose=false"],
            input=prompt,
            text=True,
            capture_output=True,
            timeout=300
        )
        return r.stdout.strip()
    except subprocess.TimeoutExpired:
        return ""

def azure_json(model, text, api_version):

    json_template = "{\n"
    for field in FIELDS:
        json_template += f' "{field}": "",\n'
    json_template = json_template.rstrip(",\n") + "\n}"

    prompt = f"""
{USER_PROMPT}

Extract the information and return ONLY a JSON object.

Rules:
- Return ONLY JSON
- No explanation
- No text before or after JSON
- If a field is missing return ""

JSON format:
{json_template}


Document:
{text}
"""

    try:
        client = AzureOpenAI(
            api_key=AZURE_API_KEY,
            azure_endpoint=AZURE_ENDPOINT,
            api_version=api_version
        )

        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.7,
            max_tokens=1500
        )

        return response.choices[0].message.content

    except Exception as e:
        print("\n----- AZURE ERROR -----")
        print("Model:", model)
        print("Error:", e)
        print("-----------------------\n")
        return ""
# ---------------- SAFE JSON PARSER ----------------
def parse_json(raw):
    try:
        raw = raw.strip()
        raw = raw.replace("```json", "").replace("```", "")

        start = raw.find("{")
        end = raw.rfind("}")

        if start == -1:
            return {}

        if end == -1:
            raw = raw + "}"
            end = raw.rfind("}")

        json_str = raw[start:end+1]
        obj = json.loads(json_str)

        # normalize keys to lowercase
        obj = {k.lower(): v for k, v in obj.items()}

        return obj

    except:
        return {}


# ---------------- MAIN ----------------
writer = pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl")
summary = []

for file in input_files:

    file = file.strip()


    print("\nProcessing:", file)

    # Detect file type
    if file.endswith(".pdf"):
        text = pdf_text(file)

    elif file.endswith(".txt"):
        text = txt_text(file)

    elif file.endswith(".docx"):
        text = docx_text(file)

    elif file.endswith(".xlsx"):
        text = excel_text(file)

    else:
        print("Unsupported file format:", file)
        continue
    
    # LLAMA
    print("Running LLAMA model...")

    llama_start = time.time()
    llama_raw = ollama_json("llama3:8b", text)
    print("LLAMA RAW:", llama_raw)

    llama = parse_json(llama_raw)
    llama_time = round(time.time() - llama_start, 2)

    print("LLAMA completed in", llama_time, "seconds")

    # MISTRAL
    print("Running MISTRAL model...")

    mistral_start = time.time()
    mistral_raw = ollama_json("mistral:7b", text)
    print("MISTRAL RAW:", mistral_raw)

    mistral = parse_json(mistral_raw)
    mistral_time = round(time.time() - mistral_start, 2)

    print("MISTRAL completed in", mistral_time, "seconds")

    # QWEN
    print("Running QWEN model...")

    qwen_start = time.time()
    qwen_raw = ollama_json("qwen2.5:7b", text)
    print("QWEN RAW:", qwen_raw)

    qwen = parse_json(qwen_raw)
    qwen_time = round(time.time() - qwen_start, 2)

    print("QWEN completed in", qwen_time, "seconds")

    # GPT-4.1 (Azure)
    print("Running GPT-4.1 Azure model...")

    gpt_start = time.time()
    gpt_raw = azure_json("gpt-4.1", text, AZURE_API_VERSION_GPT)
    print("GPT RAW:", gpt_raw)

    gpt = parse_json(gpt_raw)
    gpt_time = round(time.time() - gpt_start, 2)

    print("GPT completed in", gpt_time, "seconds")

    # Azure Llama
    print("Running AZURE LLAMA model...")

    azure_llama_start = time.time()
    azure_llama_raw = azure_json(
        "Meta-Llama-3.1-8B-Instruct",
        text,
        AZURE_API_VERSION_LLAMA
    )

    print("AZURE LLAMA RAW:", azure_llama_raw)

    azure_llama = parse_json(azure_llama_raw)
    azure_llama_time = round(time.time() - azure_llama_start, 2)

    print("AZURE LLAMA completed in", azure_llama_time, "seconds")

    # CREATE DATAFRAME
    df = pd.DataFrame({
        "Field": FIELDS,
        "Ground Truth": [""] * len(FIELDS),
        "Llama": [llama.get(x, "") for x in FIELDS],
        "Mistral": [mistral.get(x, "") for x in FIELDS],
        "Qwen": [qwen.get(x, "") for x in FIELDS],
        "GPT-4.1": [gpt.get(x, "") for x in FIELDS],
        "Azure-Llama": [azure_llama.get(x, "") for x in FIELDS],
    })

    total_row = pd.DataFrame([{
        "Field": "TOTAL_MODEL_EXECUTION_TIME (seconds)",
        "Ground Truth": "",
        "Llama": llama_time,
        "Mistral": mistral_time,
        "Qwen": qwen_time,
        "GPT-4.1": gpt_time,
        "Azure-Llama": azure_llama_time,
    }])

    df = pd.concat([df, total_row], ignore_index=True)

    sheet = os.path.basename(file)
    sheet = sheet.replace(".pdf", "").replace(".txt", "")
    sheet = re.sub(r'[\\/*?:\[\]]', "_", sheet)
    sheet = sheet[:31]

    df.to_excel(writer, sheet_name=sheet, index=False)

    # ===== FORMATTING (UNCHANGED) =====
    ws = writer.sheets[sheet]

    bold_font = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = wrap_align

    # Force TEXT format for all rows except execution time
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.number_format = "@"

    # Bold Field column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=1):
        for cell in row:
            cell.font = bold_font

    # Bold execution row
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if str(row[0].value).startswith("TOTAL_MODEL_EXECUTION_TIME"):
            for cell in row:
                cell.font = bold_font
                cell.alignment = center_align

    # Auto width
    for col in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    summary.append([sheet, llama_time, mistral_time, qwen_time, gpt_time, azure_llama_time])


summary_df = pd.DataFrame(
    summary,
    columns=["Document", "Llama_Time", "Mistral_Time", "Qwen_Time", "GPT_Time", "Azure_Llama_Time"]
)

summary_df.to_excel(writer, sheet_name="Final_Summary", index=False)
# ---------------- FINAL SUMMARY FORMATTING ----------------
ws = writer.sheets["Final_Summary"]

bold_font = Font(bold=True)
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")

center_align = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header styling
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

# Data cells formatting
for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                        min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = border
        cell.alignment = center_align

# First column bold (Document/Resume name)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                        min_col=1, max_col=1):
    for cell in row:
        cell.font = bold_font
        cell.alignment = center_align

# Force text format for document names only
for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                        min_col=1, max_col=1):
    for cell in row:
        cell.number_format = "@"

# Auto column width
for col in range(1, ws.max_column + 1):
    max_length = 0
    col_letter = get_column_letter(col)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=col, max_col=col):
        for cell in row:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

    ws.column_dimensions[col_letter].width = min(max_length + 4, 30)
writer.close()

print("\nDONE → benchmark_output.xlsx created")